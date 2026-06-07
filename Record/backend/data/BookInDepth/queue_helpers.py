"""Helpers for the autonomous queue scan.

Usage (from Claude in the loop)::

    from queue_helpers import read_pending, mark_in_progress, mark_done, mark_failed

    pending = read_pending()
    for row_idx, item in pending:
        mark_in_progress(row_idx)
        try:
            # ... generate book.json with the deep-read ...
            book_id = "some-slug"
            mark_done(row_idx, book_id=book_id)
        except Exception as e:
            mark_failed(row_idx, str(e))
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl  # type: ignore

HERE = Path(__file__).parent
QUEUE_FILE = HERE / "books_queue.xlsx"

HEADER_ROW = 1
# 1-indexed column numbers — match the header in books_queue.xlsx
COL = {
    "seq": 1,
    "title": 2,
    "author": 3,
    "originalTitle": 4,
    "year": 5,
    "status": 6,
    "addedAt": 7,
    "completedAt": 8,
    "bookId": 9,
    "notes": 10,
}


def _today() -> str:
    return datetime.now().strftime("%Y-%m-%d")


def _open_wb():
    return openpyxl.load_workbook(QUEUE_FILE)


def read_pending() -> list[tuple[int, dict[str, Any]]]:
    """Return [(row_index, row_dict), ...] for every status=='pending' row."""
    wb = _open_wb()
    ws = wb.active
    out: list[tuple[int, dict[str, Any]]] = []
    for row_idx in range(HEADER_ROW + 1, ws.max_row + 1):
        status = ws.cell(row=row_idx, column=COL["status"]).value
        if (status or "").strip().lower() != "pending":
            continue
        title = ws.cell(row=row_idx, column=COL["title"]).value
        if not title:
            continue
        out.append((row_idx, {
            "seq": ws.cell(row=row_idx, column=COL["seq"]).value,
            "title": str(title).strip(),
            "author": (ws.cell(row=row_idx, column=COL["author"]).value or "").strip(),
            "originalTitle": (ws.cell(row=row_idx, column=COL["originalTitle"]).value or "").strip(),
            "year": (ws.cell(row=row_idx, column=COL["year"]).value or "").strip(),
            "notes": (ws.cell(row=row_idx, column=COL["notes"]).value or "").strip(),
        }))
    wb.close()
    return out


def read_done_titles() -> set[str]:
    """Set of titles already marked done — used to skip duplicates."""
    wb = _open_wb()
    ws = wb.active
    out: set[str] = set()
    for row_idx in range(HEADER_ROW + 1, ws.max_row + 1):
        status = ws.cell(row=row_idx, column=COL["status"]).value
        if (status or "").strip().lower() == "done":
            title = ws.cell(row=row_idx, column=COL["title"]).value
            if title:
                out.add(str(title).strip())
    wb.close()
    return out


def mark_in_progress(row_idx: int) -> None:
    wb = _open_wb()
    ws = wb.active
    ws.cell(row=row_idx, column=COL["status"], value="in_progress")
    if not ws.cell(row=row_idx, column=COL["addedAt"]).value:
        ws.cell(row=row_idx, column=COL["addedAt"], value=_today())
    wb.save(QUEUE_FILE)
    wb.close()


def mark_done(row_idx: int, book_id: str) -> None:
    wb = _open_wb()
    ws = wb.active
    ws.cell(row=row_idx, column=COL["status"], value="done")
    ws.cell(row=row_idx, column=COL["completedAt"], value=_today())
    ws.cell(row=row_idx, column=COL["bookId"], value=book_id)
    wb.save(QUEUE_FILE)
    wb.close()


def mark_failed(row_idx: int, reason: str) -> None:
    wb = _open_wb()
    ws = wb.active
    ws.cell(row=row_idx, column=COL["status"], value="failed")
    existing_notes = ws.cell(row=row_idx, column=COL["notes"]).value or ""
    combined = f"{existing_notes} | FAIL {_today()}: {reason}".strip(" |")
    ws.cell(row=row_idx, column=COL["notes"], value=combined[:300])
    wb.save(QUEUE_FILE)
    wb.close()


def mark_skip(row_idx: int, reason: str = "") -> None:
    wb = _open_wb()
    ws = wb.active
    ws.cell(row=row_idx, column=COL["status"], value="skip")
    if reason:
        existing_notes = ws.cell(row=row_idx, column=COL["notes"]).value or ""
        combined = f"{existing_notes} | SKIP: {reason}".strip(" |")
        ws.cell(row=row_idx, column=COL["notes"], value=combined[:300])
    wb.save(QUEUE_FILE)
    wb.close()


def add_book(title: str, author: str = "", original_title: str = "", year: str = "", notes: str = "") -> int:
    """Append a new row at the bottom — useful for scripted adds.

    Returns the new row index. The user normally edits the xlsx directly in
    Excel, but this is here for batch / scripted additions.
    """
    wb = _open_wb()
    ws = wb.active
    new_row = ws.max_row + 1
    # Compute next seq
    seq = 1
    if ws.max_row > HEADER_ROW:
        prev_seq = ws.cell(row=ws.max_row, column=COL["seq"]).value
        try:
            seq = int(prev_seq) + 1 if prev_seq else 1
        except (ValueError, TypeError):
            seq = ws.max_row - HEADER_ROW + 1
    ws.cell(row=new_row, column=COL["seq"], value=seq)
    ws.cell(row=new_row, column=COL["title"], value=title)
    ws.cell(row=new_row, column=COL["author"], value=author)
    ws.cell(row=new_row, column=COL["originalTitle"], value=original_title)
    ws.cell(row=new_row, column=COL["year"], value=year)
    ws.cell(row=new_row, column=COL["status"], value="pending")
    ws.cell(row=new_row, column=COL["addedAt"], value=_today())
    ws.cell(row=new_row, column=COL["notes"], value=notes)
    wb.save(QUEUE_FILE)
    wb.close()
    return new_row


if __name__ == "__main__":
    # Quick CLI: print summary
    import sys
    pending = read_pending()
    print(f"Queue file: {QUEUE_FILE}")
    print(f"Pending: {len(pending)}")
    for row_idx, item in pending:
        print(f"  Row {row_idx}: {item['title']} (作者: {item['author'] or '?'})")
    done_count = len(read_done_titles())
    print(f"Done: {done_count}")
